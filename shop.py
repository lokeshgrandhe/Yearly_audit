from flask import Flask, render_template, request, flash, session

app = Flask(__name__)

import pandas as pd
import xlsxwriter
import openpyxl
import os, errno
import datetime
import plotly.express as px

app.secret_key = "DLF"

###### Home page #######
@app.route('/')
def s_main():
    session.clear()
    return render_template('shop_main.html')

###### Checking whether ecel exists or not #######
@app.route('/status', methods = ['POST', 'GET'])
def f_name():
    if request.method == 'POST':

        if bool(session.get('f_name', None)):
            f_n = session.get('f_name', None)
        else:
            file_name = request.form['name']
            if file_name == '':
                flash("Filed should not be empty")
                return render_template('shop_main.html')

            f_name = "./" + file_name + ".xlsx"
            f_n = file_name + ".xlsx"
            session['f_name'] = f_n

        f_path = 'C:\\Users\\Hello\\Desktop\\loki\\python\\pycharm\\shop\\Shop\\files\\'+f_n
        session['f_p'] = f_path

        if os.path.exists(f_path):
            flash("File exist")
            return render_template('exist.html', name=f_n)

        else:
            flash("Creating file")
            return render_template('create.html', name=f_n)
    else:
        f_n = session.get('f_name', None)
        flash("is selected")
        return render_template('exist.html', name=f_n)

###### Filtering to Add data or to search data #######
@app.route('/data', methods = ['POST','GET'])
def data():
    f_n = session.get('f_name', None)
    file_path = session.get('f_p', None)

    date_df = pd.read_excel(file_path)
    lis=date_df['DATE'].tolist()
    length=len(lis)

    session['len'] = length
    session['list'] = lis

    s_d = request.form['s_date']
    e_d = request.form['e_date']
    day = request.form['day']

    def date_fil(s_date, e_date, lis, date_df):

        global s_index, e_index

        if s_date in lis:
            s_index = list(date_df.index[date_df['DATE'] == s_date])
        else:
            flash("Given start date is not exist in file")
            return render_template('exist.html')

        if e_date in lis:
            e_index = list(date_df.index[date_df['DATE'] == e_date])
        else:
            flash("Given end date is not exist in file")
            return render_template('exist.html')

        date_filter_df = date_df[s_index[0]:e_index[0] + 1]
        return date_filter_df

    def silentremove(html_name):
        try:
            os.remove(html_name)
        except OSError as e:  # this would be "except OSError, e:" before Python 2.6
            if e.errno != errno.ENOENT:  # errno.ENOENT = no such file or directory
                raise

    def graph(graph):
        graph1 = graph.fillna(0)

        graph1["DATE"] = graph1["DATE"].astype("datetime64")
        ddf = graph1[['DATE', 'DAY', 'Amount', 'Expences', 'Total Amount']]

        df_date = ddf.groupby([ddf["DATE"].dt.month, ddf["DATE"].dt.year]).sum()
        df_e = df_date["Expences"]
        df_e = df_e.to_frame()
        df_e['DDD'] = df_e.index
        df_e['DDD'] = df_e['DDD'].astype(str)
        df_m = df_date
        df_m['DDD'] = df_m.index
        df_m['DDD'] = df_m['DDD'].astype(str)
        df_d = ddf.groupby([ddf["DATE"].dt.day]).sum()
        df_d['DDD'] = df_d.index
        df_dn = ddf.groupby([ddf["DAY"]]).sum()
        df_dn['DDD'] = df_dn.index
        df_e['DDD'] = df_e['DDD'].astype(str)

        html_name = r"C:\\Users\\Hello\\Desktop\\loki\\python\\pycharm\\shop\\Shop\\static\\plot.html"
        silentremove(html_name)

        fig = px.bar(x=ddf['DATE'], y=[ddf['Amount'], ddf['Total Amount'], ddf['Expences']])
        fig['data'][0]['name'] = 'Amount'
        fig['data'][1]['name'] = 'Total Amount'
        fig['data'][2]['name'] = 'Expenses'
        fig.update_layout(autosize=False, width=1350, height=350, template='plotly_white',
                          title='Amount for selected period with respect to days',
                          xaxis_title="Date", yaxis_title="Amount", yaxis_tickformat=',.')

        fig1 = px.line(x=df_e['DDD'], y=df_e['Expences'])
        fig1['data'][0]['name'] = 'Expenses'
        fig1.update_layout(autosize=False, width=1350, height=350, template='plotly_white',
                           title='Total expenses for selected period with respect to month and year',
                           xaxis_title="Month-Year", yaxis_title="Amount", yaxis_tickformat=',.')

        fig2 = px.line(x=df_m['DDD'], y=[df_m['Total Amount'], df_m['Amount'], df_m['Expences']])
        fig2['data'][0]['name'] = 'Total Amount'
        fig2['data'][1]['name'] = 'Amount'
        fig2['data'][2]['name'] = 'Expenses'
        fig2.update_layout(autosize=False, width=1350, height=350, template='plotly_white',
                           title='Total Amount details for selected period with respect to month and year',
                           xaxis_title="Month-Year", yaxis_title="Amount", yaxis_tickformat=',.')

        fig3 = px.bar(x=df_d['DDD'], y=[df_d['Total Amount'], df_d['Amount'], df_d['Expences']])
        fig3['data'][0]['name'] = 'Total Amount'
        fig3['data'][1]['name'] = 'Amount'
        fig3['data'][2]['name'] = 'Expenses'
        fig3.update_layout(autosize=False, width=1350, height=350, template='plotly_white',
                           title='Total Amount details for selected period with respect to day\'s',
                           xaxis_title="Day", yaxis_title="Amount", yaxis_tickformat=',.')

        fig4 = px.bar(x=df_dn['DDD'], y=[df_dn['Total Amount'], df_dn['Amount'], df_dn['Expences']])
        fig4['data'][0]['name'] = 'Total Amount'
        fig4['data'][1]['name'] = 'Amount'
        fig4['data'][2]['name'] = 'Expenses'
        fig4.update_layout(autosize=False, width=1350, height=350, template='plotly_white',
                           title='Total Amount details for selected period with respect to day names',
                           xaxis_title="", yaxis_title="Amount", yaxis_tickformat=',.')

        with open(r'c:\Users\Hello\Desktop\loki\python\pycharm\shop\Shop\static\plot.html', 'a') as f:
            f.write(fig.to_html(full_html=False, include_plotlyjs='cdn'))
            f.write(fig1.to_html(full_html=False, include_plotlyjs='cdn'))
            f.write(fig2.to_html(full_html=False, include_plotlyjs='cdn'))
            f.write(fig3.to_html(full_html=False, include_plotlyjs='cdn'))
            f.write(fig4.to_html(full_html=False, include_plotlyjs='cdn'))

    if not(bool(request.form.get('opt'))):
        flash("Please select atleast one option")
        return render_template('exist.html')

    if request.form['opt'] == 'add':
        return render_template('add_date.html', name = f_n)

    elif request.form['opt'] == 'all':
        lis1 = list(date_df['DATE'].dropna())
        end = lis1[-1]
        ind = list(date_df.index[date_df['DATE'] == end])
        tot_fil = date_df[:ind[0]+1]
        t_f = tot_fil[['Amount', 'Expences', 'Total Amount']].sum()
        tol_df = t_f.to_frame()
        graph(tot_fil)
        flash("For Entire Period in selected file")
        return render_template('output.html', out=[tot_fil.to_html(classes='data', index=False, na_rep='')],
                               tol=[tol_df.to_html(classes='data', header=False, na_rep='')])

    else:
        chk = request.form.getlist('chk')

        if 'chkdate' in chk and 'chkday' in chk:

            if s_d == '' or e_d == '' or day == '':
                flash("Date/Day field entered is empty")
                return render_template('exist.html')

            d = datetime.datetime.strptime(s_d, '%Y-%m-%d')
            s_date = datetime.date.strftime(d, "%d-%B-%Y")
            d1 = datetime.datetime.strptime(e_d, '%Y-%m-%d')
            e_date = datetime.date.strftime(d1, "%d-%B-%Y")

            date_filter_df = date_fil(s_date, e_date, lis, date_df)

            date_day_fil_df = date_filter_df[date_filter_df['DAY'] == day]
            d_d_t = date_day_fil_df[['Amount', 'Expences', 'Total Amount']].sum()
            date_day_tol = d_d_t.to_frame()
            graph(date_day_fil_df)
            return render_template('output.html', out=[date_day_fil_df.to_html(classes='data', index=False, na_rep='')],
                                   tol=[date_day_tol.to_html(classes='data', header=False, na_rep='')])

        elif 'chkday' in chk:
            if day == '':
                flash("Day field entered is empty")
                return render_template('exist.html')

            day_df = date_df[date_df['DAY'] == day]
            da_tol = day_df[['Amount', 'Expences', 'Total Amount']].sum()
            day_tol = da_tol.to_frame()
            graph(day_df)
            return render_template('output.html', out=[day_df.to_html(classes='data', index=False, na_rep='')],
                                   tol=[day_tol.to_html(classes='data', header=False, na_rep='')])

        elif 'chkdate' in chk:

            if s_d == '' or e_d == '':
                flash("Date field entered is empty")
                return render_template('exist.html')

            d = datetime.datetime.strptime(s_d, '%Y-%m-%d')
            s_date = datetime.date.strftime(d, "%d-%B-%Y")
            d1 = datetime.datetime.strptime(e_d, '%Y-%m-%d')
            e_date = datetime.date.strftime(d1, "%d-%B-%Y")

            date_filter_df = date_fil(s_date, e_date, lis, date_df)
            date_tol1 = date_filter_df[['Amount', 'Expences', 'Total Amount']].sum()
            date_tol = date_tol1.to_frame()
            graph(date_filter_df)
            return render_template('output.html', out=[date_filter_df.to_html(classes='data', index=False, na_rep='')],
                                   tol=[date_tol.to_html(classes='data', header=False, na_rep='')])

        else:
            flash("Kindly enter valid details")
            return render_template('exist.html')

###### Taking date from user to add data #######
@app.route('/add', methods = ['POST','GET'])
def add_data():

    lis = session.get('list', None)
    f_n = session.get('f_name', None)

    if not(bool(request.form.get('add_date'))) or request.form['add_date'] == '':
        flash("\n\nProvide value")
        return render_template('add_date.html', file=f_n)

    g_d = request.form['add_date']
    d = datetime.datetime.strptime(g_d, '%Y-%m-%d')
    given_date = datetime.date.strftime(d, "%d-%B-%Y")

    session['g_d'] = given_date

    if given_date in lis:
        flash("Add below details for ")
        return render_template('add_details.html', dat = given_date, file=f_n)

    else:
        flash("file doesn't contain given date, Please provide correct date")
        return render_template('add_date.html', err = given_date, file = f_n)

###### Adding data in excel #######
@app.route('/amount_details', methods = ['POST','GET'])
def amt_details():
    amt = request.form['amt']
    exp = request.form['exp']

    if exp == '':
        exp = 0

    while (True):
        try:
            amt = float(amt)
            exp = float(exp)
            break
        except ValueError:
            flash("Enter only numbers in Amount and Expense fields")
            return render_template('add_details.html')

    if amt <= 0 or exp < 0:
        flash("Amount should grater than 0, Enter valid amount")
        return render_template('add_details.html')

    tol = amt - exp

    lis = session.get('list', None)
    f_n = session.get('f_name', None)
    given_date = session.get('g_d', None)
    index = lis.index(given_date)
    row = index+2

    workbook = openpyxl.load_workbook(f_n)
    worksheet = workbook["My sheet"]
    tol = amt - exp
    worksheet.cell(row=row, column=3).value = amt
    worksheet.cell(row=row, column=4).value = exp
    worksheet.cell(row=row, column=5).value = tol

    worksheet.cell(row=372, column=3).value = '=SUM(C2:C367)'
    worksheet.cell(row=372, column=4).value = '=SUM(D2:D367)'
    worksheet.cell(row=372, column=5).value = '=SUM(E2:E367)'

    worksheet.cell(row=row, column=3).number_format = '#,##0.00'
    worksheet.cell(row=row, column=4).number_format = '#,##0.00'
    worksheet.cell(row=row, column=5).number_format = '#,##0.00'
    worksheet.cell(row=372, column=3).number_format = '#,##0.00'
    worksheet.cell(row=372, column=4).number_format = '#,##0.00'
    worksheet.cell(row=372, column=5).number_format = '#,##0.00'

    worksheet.freeze_panes = 'B2'
    worksheet.column_dimensions['A'].width = 20
    worksheet.column_dimensions['B'].width = 11
    worksheet.column_dimensions['C'].width = 13
    worksheet.column_dimensions['D'].width = 14
    worksheet.column_dimensions['E'].width = 14

    workbook.save(f_n)
    workbook.close

    file_path = session.get('f_p', None)
    d_df = pd.read_excel (file_path)
    df1 = d_df[['Amount','Expences','Total Amount']].apply(pd.to_numeric, errors='coerce').sum()
    df = df1.to_frame()
    return render_template('details_success.html', amt=amt, exp=exp, tot=tol, date= given_date, file=f_n,
                           data=[df.to_html(classes='date', header=False, na_rep='')])

###### Adding excel #######
@app.route('/create_new', methods = ['POST','GET'])
def c_new():
    f_n = session.get('f_name', None)
    year = request.form['s_year']
    y1 = int(year) + 1
    y2 = str(y1)
    start_year = '01-April-' + year
    end_year = '31-March-' + y2

    s1 = datetime.datetime.strptime(start_year, '%d-%B-%Y')
    s2 = datetime.datetime.strptime(end_year, '%d-%B-%Y')

    workbook = xlsxwriter.Workbook(f_n)
    worksheet = workbook.add_worksheet("My sheet")
    head_format = workbook.add_format({'bold': True, 'bg_color': '#C0C0C0'})
    day_format = workbook.add_format({'bold': True, 'bg_color': '#FFFF00'})
    head_format1 = workbook.add_format({'bold': True, 'bg_color': 'green'})
    head_format2 = workbook.add_format({'bold': True, 'bg_color': 'red'})
    head_format3 = workbook.add_format({'bold': True, 'bg_color': 'orange'})

    row = 1
    worksheet.write(0, 0, 'DATE', head_format)
    worksheet.write(0, 1, 'DAY', head_format)
    worksheet.write(0, 2, 'Amount', head_format)
    worksheet.write(0, 3, 'Expences', head_format)
    worksheet.write(0, 4, 'Total Amount', head_format)
    worksheet.write(369, 2, 'Total Amount', head_format1)
    worksheet.write(369, 3, 'Total Expences', head_format2)
    worksheet.write(369, 4, 'Grand Amount', head_format1)

    dd = pd.date_range(start=s1, end=s2).strftime('%d-%B-%Y %A')

    for item in dd:
        ddd = item.split()
        worksheet.write(row, 0, ddd[0])
        if ddd[1].lower() == 'sunday':
            worksheet.write(row, 1, ddd[1], day_format)
        else:
            worksheet.write(row, 1, ddd[1])

        row += 1
    workbook.close()

    flash("file has been created")
    return render_template('shop_main.html', f_name=f_n)

###### End of Adding excel #######

if __name__ == '__main__':
   app.run(debug = True)